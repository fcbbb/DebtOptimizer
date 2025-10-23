# core/utils/excel_export.py
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from ..models import Customer, CreditCard, Loan, MonthlyPayment, CreditCardPayment
from datetime import datetime

def export_customer_to_excel(customer):
    """
    导出客户数据为Excel，参考用户详情页排版样式
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "债务优化客户信息表"

    # --- 设置全局样式 ---
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 颜色定义
    COLORS = {
        'primary': '007bff',
        'success': '28a745',
        'info': '17a2b8',
        'warning': 'ffc107',
        'danger': 'dc3545',
        'purple': '6610f2',
        'light_blue': 'e6f2ff',
        'light_green': 'e6ffe6',
        'light_purple': 'f2e6ff',
        'light_yellow': 'fff9e6',
        'light_red': 'ffe6e6',
    }

    # --- 标题 ---
    ws['A1'] = "债务优化客户信息表"
    ws.merge_cells('A1:U1')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws['A1'].font = Font(bold=True, size=16, color=COLORS['primary'])
    ws.row_dimensions[1].height = 30

    # --- 基本信息卡片 ---
    # 标题
    ws['A3'] = "基本信息"
    ws.merge_cells('A3:U3')
    ws['A3'].font = Font(bold=True, size=12)
    ws['A3'].fill = PatternFill(start_color=COLORS['light_blue'], end_color=COLORS['light_blue'], fill_type='solid')
    ws['A3'].border = thin_border
    ws.row_dimensions[3].height = 25

    # 内容
    ws['A4'] = "客户姓名"
    ws['A4'].font = Font(bold=True)
    ws['B4'] = customer.name
    ws['B4'].alignment = Alignment(horizontal='left')

    ws['D4'] = "电话及微信号"
    ws['D4'].font = Font(bold=True)
    ws['E4'] = customer.phone
    ws['E4'].alignment = Alignment(horizontal='left')

    ws['G4'] = "融资日期"
    ws['G4'].font = Font(bold=True)
    ws['H4'] = customer.financing_date.strftime('%Y-%m-%d') if customer.financing_date else "未设置"
    ws['H4'].alignment = Alignment(horizontal='left')

    ws['J4'] = "打印时间"
    ws['J4'].font = Font(bold=True)
    ws['K4'] = datetime.now().strftime('%Y-%m-%d')
    ws['K4'].alignment = Alignment(horizontal='left')

    ws['M4'] = "信用卡倍率"
    ws['M4'].font = Font(bold=True)
    ws['N4'] = float(customer.credit_card_multiplier) if customer.credit_card_multiplier is not None and float(customer.credit_card_multiplier) != 0 else 0.03
    ws['N4'].alignment = Alignment(horizontal='left')
    ws['N4'].number_format = '0.00'  # 设置数字格式以正确显示小数

    ws['P4'] = "月供倍率"
    ws['P4'].font = Font(bold=True)
    ws['Q4'] = float(customer.monthly_payment_multiplier) if customer.monthly_payment_multiplier is not None and float(customer.monthly_payment_multiplier) != 0 else 0.002
    ws['Q4'].alignment = Alignment(horizontal='left')
    ws['Q4'].number_format = '0.000'  # 设置数字格式以正确显示小数

    ws['S4'] = "备注"
    ws['S4'].font = Font(bold=True)
    ws['T4'] = customer.notes if customer.notes else "无"
    ws.merge_cells('T4:U4')
    ws['T4'].alignment = Alignment(horizontal='left')
    
    # 第二行展示信用卡数量和贷款数量
    ws['A5'] = "信用卡数量"
    ws['A5'].font = Font(bold=True)
    ws['B5'] = customer.creditcard_set.all().count()
    ws['B5'].alignment = Alignment(horizontal='left')
    
    ws['D5'] = "贷款数量"
    ws['D5'].font = Font(bold=True)
    ws['E5'] = customer.loan_set.all().count()
    ws['E5'].alignment = Alignment(horizontal='left')

    # 设置边框
    for row in [4, 5]:
        for col in ['A', 'B', 'D', 'E', 'G', 'H', 'J', 'K', 'M', 'N', 'P', 'Q', 'S', 'T']:
            if row == 5 and col in ['G', 'H', 'J', 'K', 'M', 'N', 'P', 'Q']:
                continue  # 第二行没有这些字段
            ws[f'{col}{row}'].border = thin_border
    ws['U4'].border = thin_border
    ws['U5'].border = thin_border

    # --- 财务概览卡片 ---
    # 标题
    ws['A6'] = "财务概览"
    ws.merge_cells('A6:U6')
    ws['A6'].font = Font(bold=True, size=12)
    ws['A6'].fill = PatternFill(start_color=COLORS['light_purple'], end_color=COLORS['light_purple'], fill_type='solid')
    ws['A6'].border = thin_border
    ws.row_dimensions[6].height = 25

    # 统计卡片
    # 计算0账单费用总计（信用卡费率3%总和）
    total_credit_card_fee = sum(float(cp.fee) if cp.fee is not None else 0 for cp in customer.creditcardpayment_set.all()) or 0
    
    stats = [
        {'title': '总负债', 'value': customer.total_debt or 0, 'color': 'light_blue', 'pos': 'A7'},
        {'title': '总月供', 'value': customer.total_monthly_payment or 0, 'color': 'light_green', 'pos': 'E7'},
        {'title': '总出款', 'value': customer.total_payment or 0, 'color': 'light_purple', 'pos': 'I7'},
        {'title': '总使用成本', 'value': sum(p.cost for p in customer.monthlypayment_set.all() if p.cost) or 0, 'color': 'light_yellow', 'pos': 'M7'},
        {'title': '本金成本总计', 'value': (sum(p.cost for p in customer.monthlypayment_set.all() if p.cost) or 0) + (customer.total_payment or 0), 'color': 'light_red', 'pos': 'Q7'},
        {'title': '0账单费用总计', 'value': total_credit_card_fee, 'color': 'light_red', 'pos': 'U7'},
    ]

    for stat in stats:
        # 标题
        title_cell = ws[stat['pos']]
        title_cell.value = stat['title']
        title_cell.font = Font(bold=True)
        title_cell.alignment = Alignment(horizontal='center')
        title_cell.fill = PatternFill(start_color=COLORS[stat['color']], end_color=COLORS[stat['color']], fill_type='solid')
        title_cell.border = thin_border

        # 值
        value_cell = ws[f'{stat["pos"][0]}{int(stat["pos"][1:])+1}']
        value_cell.value = stat['value']
        value_cell.font = Font(bold=True, size=14)
        value_cell.alignment = Alignment(horizontal='center', vertical='center')
        value_cell.fill = PatternFill(start_color=COLORS[stat['color']], end_color=COLORS[stat['color']], fill_type='solid')
        value_cell.border = thin_border

        # 合并单元格
        ws.merge_cells(f'{stat["pos"]}:{chr(ord(stat["pos"][0])+3)}{int(stat["pos"][1:])}')
        ws.merge_cells(f'{stat["pos"][0]}{int(stat["pos"][1:])+1}:{chr(ord(stat["pos"][0])+3)}{int(stat["pos"][1:])+1}')

    # --- 信用卡账单明细 ---
    # 标题
    ws['A10'] = "信用卡账单明细"
    ws.merge_cells('A10:U10')
    ws['A10'].font = Font(bold=True, size=12, color=COLORS['info'])
    ws['A10'].fill = PatternFill(start_color=COLORS['light_blue'], end_color=COLORS['light_blue'], fill_type='solid')
    ws['A10'].border = thin_border
    ws.row_dimensions[10].height = 25

    # 表头
    headers_cc = ["银行渠道", "总授信额度", "有无分期", "分期金额", "账单日", "还款日"]
    for col_idx, header in enumerate(headers_cc, start=1):  # A列开始
        cell = ws.cell(row=11, column=col_idx, value=header)
        cell.border = thin_border
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color=COLORS['light_blue'], end_color=COLORS['light_blue'], fill_type='solid')
        cell.alignment = Alignment(horizontal='center')

    # 填充数据
    for idx, card in enumerate(customer.creditcard_set.all(), start=12):
        ws.cell(row=idx, column=1, value=card.bank).border = thin_border
        ws.cell(row=idx, column=2, value=float(card.total_limit) if card.total_limit is not None else 0).border = thin_border
        ws.cell(row=idx, column=3, value="是" if card.has_installment else "否").border = thin_border
        ws.cell(row=idx, column=4, value=float(card.installment_amount) if card.installment_amount is not None else 0).border = thin_border
        ws.cell(row=idx, column=5, value=card.billing_date).border = thin_border
        ws.cell(row=idx, column=6, value=card.repayment_date).border = thin_border

        # 斑马纹效果
        if idx % 2 == 0:
            for col in range(1, 7):
                ws.cell(row=idx, column=col).fill = PatternFill(start_color='f8f9fa', end_color='f8f9fa', fill_type='solid')

    # --- 信用贷款账单明细 ---
    # 计算信用卡表格后的起始行
    cc_end_row = 12 + customer.creditcard_set.all().count() + 1

    # 标题
    ws[f'A{cc_end_row}'] = "信用贷款账单明细"
    ws.merge_cells(f'A{cc_end_row}:U{cc_end_row}')
    ws[f'A{cc_end_row}'].font = Font(bold=True, size=12, color=COLORS['warning'])
    ws[f'A{cc_end_row}'].fill = PatternFill(start_color=COLORS['light_yellow'], end_color=COLORS['light_yellow'], fill_type='solid')
    ws[f'A{cc_end_row}'].border = thin_border
    ws.row_dimensions[cc_end_row].height = 25

    # 表头
    headers_loan = ["银行渠道", "总授信额度", "贷款余额", "月还款", "到期时间", "还款日"]
    for col_idx, header in enumerate(headers_loan, start=1):
        cell = ws.cell(row=cc_end_row+1, column=col_idx, value=header)
        cell.border = thin_border
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color=COLORS['light_yellow'], end_color=COLORS['light_yellow'], fill_type='solid')
        cell.alignment = Alignment(horizontal='center')

    # 填充数据
    for idx, loan in enumerate(customer.loan_set.all(), start=cc_end_row+2):
        ws.cell(row=idx, column=1, value=loan.bank).border = thin_border
        ws.cell(row=idx, column=2, value=float(loan.total_limit) if loan.total_limit is not None else 0).border = thin_border
        ws.cell(row=idx, column=3, value=float(loan.balance) if loan.balance is not None else 0).border = thin_border
        ws.cell(row=idx, column=4, value=float(loan.monthly_payment) if loan.monthly_payment is not None else 0).border = thin_border
        ws.cell(row=idx, column=5, value=loan.due_date).border = thin_border
        ws.cell(row=idx, column=6, value=loan.repayment_date).border = thin_border

        # 斑马纹效果
        if idx % 2 == 0:
            for col in range(1, 7):
                ws.cell(row=idx, column=col).fill = PatternFill(start_color='f8f9fa', end_color='f8f9fa', fill_type='solid')

    # --- 月供出款明细 ---
    # 计算贷款表格后的起始行
    loan_end_row = cc_end_row + 2 + customer.loan_set.all().count() + 1

    # 标题
    ws[f'A{loan_end_row}'] = "月供出款明细"
    ws.merge_cells(f'A{loan_end_row}:U{loan_end_row}')
    ws[f'A{loan_end_row}'].font = Font(bold=True, size=12, color=COLORS['success'])
    ws[f'A{loan_end_row}'].fill = PatternFill(start_color=COLORS['light_green'], end_color=COLORS['light_green'], fill_type='solid')
    ws[f'A{loan_end_row}'].border = thin_border
    ws.row_dimensions[loan_end_row].height = 25

    # 表头
    headers_mp = ["出款时间", "出款金额", "是否私人借款", "资金使用天数", "资金使用成本", "备注"]
    for col_idx, header in enumerate(headers_mp, start=1):
        cell = ws.cell(row=loan_end_row+1, column=col_idx, value=header)
        cell.border = thin_border
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color=COLORS['light_green'], end_color=COLORS['light_green'], fill_type='solid')
        cell.alignment = Alignment(horizontal='center')

    # 填充数据
    for idx, mp in enumerate(customer.monthlypayment_set.all(), start=loan_end_row+2):
        ws.cell(row=idx, column=1, value=mp.payment_date).border = thin_border
        ws.cell(row=idx, column=2, value=float(mp.amount) if mp.amount is not None else 0).border = thin_border
        ws.cell(row=idx, column=3, value="是" if mp.is_private else "否").border = thin_border
        ws.cell(row=idx, column=4, value=mp.days_used if mp.days_used is not None else 0).border = thin_border
        ws.cell(row=idx, column=5, value=float(mp.cost) if mp.cost is not None else 0).border = thin_border
        ws.cell(row=idx, column=6, value=mp.notes).border = thin_border

        # 斑马纹效果
        if idx % 2 == 0:
            for col in range(1, 7):
                ws.cell(row=idx, column=col).fill = PatternFill(start_color='f8f9fa', end_color='f8f9fa', fill_type='solid')

    # --- 信用卡出款明细 ---
    # 计算月供表格后的起始行
    mp_end_row = loan_end_row + 2 + customer.monthlypayment_set.all().count() + 1

    # 标题
    ws[f'A{mp_end_row}'] = "信用卡出款明细"
    ws.merge_cells(f'A{mp_end_row}:U{mp_end_row}')
    ws[f'A{mp_end_row}'].font = Font(bold=True, size=12, color=COLORS['danger'])
    ws[f'A{mp_end_row}'].fill = PatternFill(start_color=COLORS['light_red'], end_color=COLORS['light_red'], fill_type='solid')
    ws[f'A{mp_end_row}'].border = thin_border
    ws.row_dimensions[mp_end_row].height = 25

    # 表头
    headers_cp = ["出款时间", "银行", "出款金额", "信用卡费率", "刷出金额", "刷出时间", "备注"]
    for col_idx, header in enumerate(headers_cp, start=1):
        cell = ws.cell(row=mp_end_row+1, column=col_idx, value=header)
        cell.border = thin_border
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color=COLORS['light_red'], end_color=COLORS['light_red'], fill_type='solid')
        cell.alignment = Alignment(horizontal='center')

    # 填充数据
    for idx, cp in enumerate(customer.creditcardpayment_set.all(), start=mp_end_row+2):
        ws.cell(row=idx, column=1, value=cp.payment_date).border = thin_border
        ws.cell(row=idx, column=2, value=float(cp.payment_amount) if cp.payment_amount is not None else 0).border = thin_border
        ws.cell(row=idx, column=3, value=float(cp.fee) if cp.fee is not None else 0).border = thin_border
        ws.cell(row=idx, column=4, value=float(cp.withdrawal_amount) if cp.withdrawal_amount is not None else 0).border = thin_border
        ws.cell(row=idx, column=5, value=cp.withdrawal_date).border = thin_border
        ws.cell(row=idx, column=6, value=cp.bank).border = thin_border
        ws.cell(row=idx, column=7, value=cp.notes).border = thin_border

        # 斑马纹效果
        if idx % 2 == 0:
            for col in range(1, 8):
                ws.cell(row=idx, column=col).fill = PatternFill(start_color='f8f9fa', end_color='f8f9fa', fill_type='solid')

    # --- 调整列宽 ---
    column_widths = {
        'A': 15,  # 出款时间/银行渠道
        'B': 15,  # 出款金额/总授信额度
        'C': 12,  # 是否私人借款/有无分期
        'D': 12,  # 资金使用天数/分期金额
        'E': 12,  # 资金使用成本/账单日
        'F': 10,  # 备注/还款日
        'G': 15,  # 扩展列
        'H': 15,  # 扩展列
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # 设置剩余列宽
    for col in range(9, 22):  # I到U列
        ws.column_dimensions[get_column_letter(col)].width = 10

    return wb